import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border
import json

# Загружаем Excel файл
file_path = r'D:\Job\Projects\construction-management\Мастер_карта_развития_26_12_25_корр.xlsx'

# Открываем workbook для анализа форматирования
wb = openpyxl.load_workbook(file_path)

print("=" * 80)
print("АНАЛИЗ EXCEL ФАЙЛА: Мастер_карта_развития_26_12_25_корр.xlsx")
print("=" * 80)

# Список листов
print(f"\nЛисты в файле: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"ЛИСТ: {sheet_name}")
    print("="*80)
    
    ws = wb[sheet_name]
    
    # Размеры листа
    print(f"Размер: {ws.max_row} строк x {ws.max_column} столбцов")
    
    # Объединённые ячейки
    if ws.merged_cells.ranges:
        print(f"\nОбъединённые ячейки ({len(ws.merged_cells.ranges)} шт):")
        for i, merged_range in enumerate(list(ws.merged_cells.ranges)[:20]):
            print(f"  {merged_range}")
        if len(ws.merged_cells.ranges) > 20:
            print(f"  ... и ещё {len(ws.merged_cells.ranges) - 20}")
    
    # Читаем данные с помощью pandas
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        print(f"\n--- Первые 30 строк данных ---")
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 50)
        print(df.head(30).to_string())
        
        print(f"\n--- Уникальные значения в каждом столбце (первые 10) ---")
        for col in df.columns[:15]:
            unique_vals = df[col].dropna().unique()[:10]
            if len(unique_vals) > 0:
                print(f"Столбец {col}: {list(unique_vals)}")
    except Exception as e:
        print(f"Ошибка при чтении листа: {e}")
    
    # Анализ форматирования первых ячеек
    print(f"\n--- Форматирование заголовков ---")
    for row in range(1, min(5, ws.max_row + 1)):
        for col in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                fill_color = None
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    fill_color = cell.fill.fgColor.rgb
                font_color = None
                if cell.font and cell.font.color and cell.font.color.rgb:
                    font_color = cell.font.color.rgb
                print(f"  [{row},{col}] '{str(cell.value)[:30]}' - Fill: {fill_color}, Font: {font_color}, Bold: {cell.font.bold if cell.font else None}")

print("\n" + "="*80)
print("АНАЛИЗ ЗАВЕРШЁН")
print("="*80)
